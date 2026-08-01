#!/usr/bin/env python3
"""Build PROTON commercials workbook in the standard presales format
(3 tabs: GCP Cost Estimation · Timeline · PS Utilization), matching the template
'Customer Name - GCP Pricing, Timeline and PS Utilization (Format).xlsx'.

Data:
  - GCP pricing  -> PROTON_GCP_Cost_Estimate.csv (asia-southeast1, one prod tenant)
  - Timeline/PS  -> PROTON-CRM-WBS-Gantt.csv task list (4 phases, ~16 weeks)
Currency: USD + IDR (rate in an editable cell), Rp for PS — same as the template.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE = "/Users/yudaadipratama/Archive/id-crm-ticketing/docs/proposals"
OUT = f"{BASE}/PROTON - GCP Pricing, Timeline and PS Utilization.xlsx"

RATE = 16500  # USD -> IDR (editable)

# ---- palette (from template) ----
BLUE   = "FF4285F4"; BLUE2 = "FF3C78D8"; YELLOW = "FFFFF2CC"; PINK = "FFF4CCCC"
GREY1  = "FFCCCCCC"; GREY2 = "FFD9D9D9"; GREENDK = "FF274E13"; GREENLT = "FFD9EAD3"
BAR    = "FF9FC5F8"; WHITE = "FFFFFFFF"; GREYTX = "FF666666"
GS = "Google Sans"; AR = "Arial"
USD_FMT = '[$USD ]#,##0.00'; IDR_FMT = '[$IDR ]#,##0.00'
RP_FMT  = '[$Rp ]#,##0.00'; RP0_FMT = '[$Rp ]#,##0'; PCT = '0%'

thin = Side(style="thin", color="FFB7B7B7")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell(ws, r, c, v=None, *, fill=None, font_name=AR, size=10, bold=False, color=None,
         border=BOX, ha=None, va="center", wrap=False, nf=None):
    cl = ws.cell(r, c)
    if v is not None:
        cl.value = v
    cl.font = Font(name=font_name, size=size, bold=bold, color=color)
    if fill:
        cl.fill = PatternFill("solid", fgColor=fill)
    if border:
        cl.border = border
    cl.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if nf:
        cl.number_format = nf
    return cl

def box(ws, r1, c1, r2, c2, **kw):
    """Style every cell in a rect then merge (keeps full outline on merged range)."""
    v = kw.pop("v", None)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell(ws, r, c, **kw)
    if v is not None:
        ws.cell(r1, c1).value = v
    if (r1, c1) != (r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w

def fit_landscape(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.2
    ws.page_margins.top = ws.page_margins.bottom = 0.3

wb = openpyxl.Workbook()

# ============================================================ TAB 1: GCP COST
ws = wb.active
ws.title = "GCP Cost Estimation"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 5, "C": 30, "D": 52, "E": 15, "F": 17})
ws["A1"] = RATE
ws["A1"].font = Font(name=AR, size=9, color="FF999999")

def gcp_table(top, title, rows, note=None):
    r = top
    if title:
        box(ws, r, 2, r, 6, v=title, fill=GREY2, font_name=AR, bold=True, ha="left")
        r += 1
    hdr = ["No", "GCP Component", "Component Details", "Price (USD)", "Price (IDR)"]
    for i, h in enumerate(hdr):
        cell(ws, r, 2 + i, h, fill=BLUE, font_name=AR, bold=True, color=WHITE, ha="center", wrap=True)
    r += 1
    first = r
    for i, (comp, det, usd) in enumerate(rows, 1):
        cell(ws, r, 2, i, ha="center")
        cell(ws, r, 3, comp, wrap=True, ha="left")
        cell(ws, r, 4, det, wrap=True, ha="left")
        cell(ws, r, 5, usd, nf=USD_FMT, ha="center")
        cell(ws, r, 6, f"=E{r}*A$1", nf=IDR_FMT, ha="center")
        ws.row_dimensions[r].height = 14 * (1 + det.count("\n"))
        r += 1
    last = r - 1
    # totals
    box(ws, r, 2, r, 4, v="Total Price Estimation Per-Month", fill=YELLOW, bold=True, ha="right")
    cell(ws, r, 5, f"=SUM(E{first}:E{last})", fill=YELLOW, bold=True, nf=USD_FMT, ha="right")
    cell(ws, r, 6, f"=SUM(F{first}:F{last})", fill=YELLOW, bold=True, nf=IDR_FMT, ha="right")
    my = r; r += 1
    box(ws, r, 2, r, 4, v="Total Price Estimation Per-Year", fill=YELLOW, bold=True, ha="right")
    cell(ws, r, 5, f"=E{my}*12", fill=YELLOW, bold=True, nf=USD_FMT, ha="right")
    cell(ws, r, 6, f"=F{my}*12", fill=YELLOW, bold=True, nf=IDR_FMT, ha="right")
    r += 1
    box(ws, r, 2, r, 6, v="Calculator Link:  https://cloud.google.com/products/calculator",
        fill=None, ha="left")
    r += 1
    if note:
        cell(ws, r, 2, note, border=None, font_name=AR, size=9, color="FF999999", ha="left")
        r += 1
    return my, r

recurring = [
    ("Google Kubernetes Engine (Standard) — cluster",
     "Region: Singapore (asia-southeast1)\nRegional cluster management fee ($0.10/cluster/hr)\nMulti-tenant: per-tenant namespaces + node pools", 73.00),
    ("Google Kubernetes Engine (Standard) — node pool",
     "3 × e2-standard-4 (vCPUs: 4, RAM: 16 GiB), on-demand\nRuns Chatwoot (Rails+Sidekiq), agent & backend (FastAPI)", 362.01),
    ("Cloud SQL for PostgreSQL (HA)",
     "Region: Singapore (asia-southeast1)\nEnterprise edition, regional HA\nMachine: 2 vCPU / 8 GiB RAM\nStorage: 100 GB SSD\nChatwoot DB + per-tenant pgvector KB", 300.00),
    ("Memorystore for Redis (HA)",
     "Standard tier (auto-failover replica)\nCapacity: 4 GB\nSidekiq queues / cache", 181.04),
    ("Vertex AI — Gemini 2.5 Flash + Vertex AI Search",
     "Usage-based (indicative)\n~5,000 AI conversations/mo (~8k in + ~1.5k out tokens)\n~5,000 grounded KB queries/mo", 50.75),
    ("Cloud Storage",
     "Region: Singapore (asia-southeast1)\nClass: Standard · 100 GB + modest egress\nAttachments / exports", 8.30),
    ("Networking — Cloud Load Balancing + egress",
     "1 external HTTP(S) forwarding rule + data processing\n~100 GB internet egress/mo", 31.85),
    ("Cloud Logging + Cloud Monitoring",
     "~100 GB log ingest/mo (first 50 GB free)\nMetrics + alerting", 25.00),
    ("Secret Manager + Artifact Registry + Cloud Build + BigQuery",
     "Secrets · container images · CI builds · analytics warehouse (storage + query)", 17.25),
]
onetime = [
    ("Cloud Storage — migration staging",
     "Zendesk export staging bucket (Standard, ~50 GB, short-lived)", 5.00),
    ("Data transfer / egress — migration",
     "One-time import/export bandwidth for tickets, contacts, KB, attachments", 15.00),
    ("Temporary migration compute",
     "Short-lived job for bulk import + KB ingest into Vertex AI Search / pgvector", 20.00),
]

box(ws, 2, 2, 2, 6, v=None)  # placeholder (styled by table header below)
my1, nxt = gcp_table(2, "PROTON CRM — Monthly Recurring (one production tenant)", recurring)
_, nxt2 = gcp_table(nxt + 1, "One-time — Migration Period *", onetime,
                    note="*one time only during migration process")
box(ws, nxt2 + 1, 2, nxt2 + 1, 6,
    v=f"IDR shown at USD→IDR = cell A1 ({RATE:,}, editable). Indicative GCP list price (Jul 2026, asia-southeast1 / "
      f"Singapore — nearest full-service region; Malaysia region not yet GA). Excludes negotiated discounts, taxes & "
      f"Google program funding. 1-yr committed-use discount lowers nodes ~$233 → ~$920/mo total.",
    fill=None, border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws.row_dimensions[nxt2 + 1].height = 42
fit_landscape(ws)

# ============================================================ TAB 2: TIMELINE
ws2 = wb.create_sheet("Timeline")
ws2.sheet_view.showGridLines = False
NW = 16
widths(ws2, {"B": 4, "C": 5, "D": 46})
for i in range(NW):
    ws2.column_dimensions[get_column_letter(5 + i)].width = 3.8

# header row
box(ws2, 2, 2, 2, 3, v="Phases", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")
cell(ws2, 2, 4, "Task", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")
for i in range(NW):
    cell(ws2, 2, 5 + i, f"W{i+1}", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")

# phases: (name, phase_no, [(subno, task, wk_start, wk_end)])
phases = [
    ("Assessment & Planning", 1, [
        ("1.1", "Kick-off, scope & SOP validation; confirm requirements from BRD + Process-Flow workbook", 1, 1),
        ("1.2", "Solution architecture, migration plan & GCP landing zone design (GKE Standard, VPC, IAM)", 1, 2),
    ]),
    ("GCP Foundation & Provisioning", 2, [
        ("2.1", "Provision GCP project, VPC, GKE Standard cluster + node pools, Cloud SQL, Memorystore, BigQuery", 2, 3),
        ("2.2", "CI/CD (Artifact Registry + Cloud Build), Secret Manager, monitoring baseline", 2, 3),
    ]),
    ("Core Platform Cutover", 3, [
        ("3.1", "Deploy Chatwoot + agent + backend on GKE Standard + Cloud SQL", 3, 4),
        ("3.2", "Zendesk data migration (tickets, contacts, KB, attachments) + parallel-run", 4, 6),
        ("3.3", "Wire channels: WhatsApp (Twilio), Website, API; KB migration to Vertex AI Search", 4, 6),
        ("3.4", "AI layer + SOP lifecycle flows (disclaimer, idle-close, YES/NO, rating, auto-classify)", 5, 7),
    ]),
    ("Gap Closure (P2 priorities)", 4, [
        ("4.1", "Email escalation — auto-ack once/thread + separate internal case (SOP-driven)", 6, 8),
        ("4.2", "Customer 360 + DMS/TSP integration (4-section card, <3s sync)", 7, 11),
        ("4.3", "Case management — hierarchical categories, per-ticket dashboard, RBAC", 8, 11),
        ("4.4", "Reporting & BI — BigQuery + Power BI, role-based views, anomaly dashboard, NPS/CRR", 9, 12),
        ("4.5", "Agent management (channel priority, auto-busy) + multimodal / language + bulk KB upload", 10, 12),
    ]),
    ("Hardening, UAT & Handover", 5, [
        ("5.1", "HA / autoscaling, monitoring & alerting, backups / DR", 12, 14),
        ("5.2", "UAT, data reconciliation & sign-off", 13, 15),
        ("5.3", "Cut-over, Zendesk decommission, ops runbook + admin/agent training", 15, 16),
    ]),
]
phase_fills = [GREY1, GREY2]
r = 3
for pi, (pname, pno, tasks) in enumerate(phases):
    # phase title row (merged B:D) + shade the week strip
    box(ws2, r, 2, r, 4, v=pname, fill=phase_fills[pi % 2], font_name=GS, bold=True, color=GREYTX, ha="center")
    for i in range(NW):
        cell(ws2, r, 5 + i, fill=phase_fills[pi % 2])
    r += 1
    start_task_row = r
    for (subno, task, ws_, we_) in tasks:
        cell(ws2, r, 3, subno, font_name=GS, color=GREYTX, ha="center")
        cell(ws2, r, 4, task, font_name=GS, color=GREYTX, ha="left", wrap=True)
        for i in range(NW):
            wk = i + 1
            if ws_ <= wk <= we_:
                cell(ws2, r, 5 + i, fill=BAR)
            else:
                cell(ws2, r, 5 + i)
        ws2.row_dimensions[r].height = 26
        r += 1
    # phase number in col B, merged across its task rows
    box(ws2, start_task_row, 2, r - 1, 2, v=f"{pno}.0", font_name=GS, bold=True, color=GREYTX, ha="center")

fit_landscape(ws2)

# ============================================================ TAB 3: PS UTILIZATION
ws3 = wb.create_sheet("PS Utilization")
ws3.sheet_view.showGridLines = False
widths(ws3, {"A": 9, "B": 24, "C": 52, "D": 20, "E": 11, "F": 18, "G": 2,
             "H": 12, "I": 15, "J": 15, "K": 15, "L": 15, "M": 15, "N": 15, "O": 16,
             "P": 2, "Q": 20, "R": 18, "S": 15})
ws3["A1"] = RATE
ws3["A1"].font = Font(name=AR, size=9, color="FF999999")

# ---- rate card (Q:S) ----
rate_rows = [
    ("Consultant", 350000),
    ("Senior Consultant", 606250),
    ("Lead Consultant", 700000),
    ("Principal", 825000),
    ("Expert / Director", 1275000),
]
cell(ws3, 2, 17, "Role / Seniority", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="left")
cell(ws3, 2, 18, "Sales price / hour (IDR)", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="center", wrap=True)
cell(ws3, 2, 19, "Per day (× 8)", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="center", wrap=True)
rr = 3
rate_cell = {}
for name, rate in rate_rows:
    cell(ws3, rr, 17, name, fill=GREENDK, font_name=GS, size=9, bold=True, color=WHITE, border=None, ha="left")
    cell(ws3, rr, 18, rate, fill=GREENLT, font_name=GS, size=9, border=None, ha="center", nf=RP0_FMT)
    cell(ws3, rr, 19, f"=R{rr}*8", fill=GREENLT, font_name=GS, size=9, border=None, ha="center", nf=RP0_FMT)
    rate_cell[name] = f"$R${rr}"
    rr += 1
cell(ws3, rr, 17, "Margin", fill=GREENDK, font_name=GS, size=9, bold=True, color=WHITE, border=None, ha="left")
cell(ws3, rr, 18, 0.15, fill=GREENLT, font_name=AR, bold=True, border=None, ha="right", nf=PCT)
MARGIN = f"$R${rr}"

# ---- role columns (I..N) : (header seniority, header function, rate_cell) ----
roles = [
    ("Lead Consultant", "Project Manager", rate_cell["Lead Consultant"]),
    ("Principal", "Solution Architect / Tech Lead", rate_cell["Principal"]),
    ("Senior Consultant", "Backend / AI Engineer", rate_cell["Senior Consultant"]),
    ("Consultant", "Frontend Engineer", rate_cell["Consultant"]),
    ("Senior Consultant", "Data / BI Engineer", rate_cell["Senior Consultant"]),
    ("Consultant", "QA Engineer", rate_cell["Consultant"]),
]
# ---- phases: (milestone, activity, resources, weeks, [md per role]) ----
ps_phases = [
    ("Phase 1 : Assessment & Planning",
     "Kick-off, scope & SOP validation; confirm requirements from BRD + Process-Flow workbook; solution architecture & migration plan.",
     "Project Manager\nSolution Architect\nBackend/AI Engineer\nData/BI Engineer", 2, [3, 6, 4, 0, 2, 1]),
    ("Phase 2 : GCP Foundation & Provisioning",
     "Provision GCP project, VPC, GKE Standard cluster + node pools, Cloud SQL, Memorystore, BigQuery; CI/CD, Secret Manager, monitoring baseline.",
     "Solution Architect\nDevOps / Backend\nData/BI Engineer", 2, [3, 6, 8, 1, 2, 1]),
    ("Phase 3 : Core Platform Cutover",
     "Deploy Chatwoot + agent + backend on GKE Standard + Cloud SQL; Zendesk data migration + parallel-run; wire WhatsApp/Web/API; KB migration; AI + SOP lifecycle flows.",
     "Project Manager\nBackend/AI Engineer\nFrontend Engineer\nData/BI Engineer\nQA", 4, [5, 5, 26, 8, 6, 6]),
    ("Phase 4 : Gap Closure (P2 priorities)",
     "Email escalation; Customer 360 + DMS/TSP; case management + hierarchical categories + RBAC; reporting & Power BI; agent management + multimodal / language + bulk KB upload.",
     "Backend/AI Engineer\nFrontend Engineer\nData/BI Engineer\nQA", 6, [6, 5, 40, 20, 16, 12]),
    ("Phase 5 : Hardening, HA & UAT",
     "HA / autoscaling, monitoring & alerting, backups / DR; UAT, data reconciliation & sign-off.",
     "DevOps / Backend\nQA\nProject Manager", 2, [4, 2, 12, 4, 2, 12]),
    ("Phase 6 : Cutover, Handover & Training",
     "Cut-over, Zendesk decommission; ops runbook + admin/agent training + knowledge transfer.",
     "Project Manager\nBackend/AI Engineer\nTechnical Writer", 2, [3, 2, 10, 3, 2, 4]),
]

# title + headers
box(ws3, 3, 2, 3, 6, v="Professional Services — PROTON CRM (self-managed on Google Cloud)",
    fill=PINK, font_name=GS, bold=True, wrap=True, ha="center")
for c, h in [(2, "Milestone"), (3, "Activity"), (4, "Resources"), (5, "Duration (Week)"), (6, "Service Price")]:
    cell(ws3, 4, c, h, fill=PINK, font_name=GS, bold=True, ha="center", wrap=True)
cell(ws3, 4, 8, "Duration (Days)", fill=YELLOW, font_name=GS, bold=True, ha="center", wrap=True)
for i, (sen, fn, _) in enumerate(roles):
    cell(ws3, 3, 9 + i, sen, fill=YELLOW, font_name=GS, size=9, bold=True, ha="center", wrap=True)
    cell(ws3, 4, 9 + i, fn, fill=YELLOW, font_name=GS, size=9, bold=True, ha="center", wrap=True)
cell(ws3, 3, 15, "Total", fill=YELLOW, font_name=GS, bold=True, ha="center")
cell(ws3, 4, 15, "(mandays)", fill=YELLOW, font_name=GS, size=9, bold=True, ha="center")

first = 5
r = first
for (mil, act, res, wks, mds) in ps_phases:
    cell(ws3, r, 2, mil, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 3, act, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 4, res, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 5, wks, font_name=GS, ha="center", nf='#,##0')
    cell(ws3, r, 8, f"=E{r}*5", font_name=GS, ha="center")
    for i, md in enumerate(mds):
        cell(ws3, r, 9 + i, md, font_name=GS, ha="center")
    cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=GS, ha="center")
    ws3.row_dimensions[r].height = 60
    r += 1
last = r - 1

# Service Price = one merged cell (F5:F10) = total incl margin (O row_total = last+3), rounded
box(ws3, first, 6, last, 6, v=f"=ROUNDUP(O{last+3},-2)", font_name=GS, bold=True, ha="center", nf=RP0_FMT)

# totals block
cell(ws3, r, 8, "Total Mandays", fill=YELLOW, font_name=AR, bold=True, ha="left")
for i in range(len(roles)):
    cl = get_column_letter(9 + i)
    cell(ws3, r, 9 + i, f"=SUM({cl}{first}:{cl}{last})", font_name=AR, ha="center")
cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=AR, bold=True, ha="center")
row_md = r; r += 1

cell(ws3, r, 8, "Total Mandays Price", fill=YELLOW, font_name=AR, bold=True, ha="left")
for i, (_, _, ratec) in enumerate(roles):
    cl = get_column_letter(9 + i)
    cell(ws3, r, 9 + i, f"={cl}{row_md}*8*{ratec}", font_name=AR, ha="center", nf=RP0_FMT)
cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=AR, bold=True, ha="center", nf=RP0_FMT)
row_price = r; r += 1

cell(ws3, r, 8, "Total Price Including Margin", fill=YELLOW, font_name=AR, bold=True, ha="left", wrap=True)
box(ws3, r, 9, r, 14, v=None)
cell(ws3, r, 15, f"=O{row_price}*(1+{MARGIN})", font_name=AR, bold=True, ha="center", nf=RP0_FMT)
row_total = r

# note
cell(ws3, r + 2, 2, "Mandays are an indicative delivery estimate; day-rates & margin are editable in the rate card (Q:S). "
                    "Rp figures use the rates above; USD equivalent at A1 rate. May be offset by Google program funding.",
     border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws3.merge_cells(start_row=r + 2, start_column=2, end_row=r + 2, end_column=8)
fit_landscape(ws3)

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
