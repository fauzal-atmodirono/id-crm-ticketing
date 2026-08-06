#!/usr/bin/env python3
"""ACCELERATED (2-month) build of the PROTON commercials workbook for the scope
OUTSIDE Packages A-G.

Same scope as `build_proton_commercials_remaining_xlsx.py` — platform delivery
that has never started, plus the product items with a spec/plan but no code
(native SAML SSO, RBAC phase 3, IVR-4, custom agent status labels,
release/enablement, live verification). Packages A-G stay excluded; they are
already queued and priced separately.

The difference is delivery method. This variant assumes the accelerator is
used, so the team parameterises pre-built assets instead of authoring them:
  - landing zone and GKE/Helm deployment come from accelerator IaC
  - Zendesk extract/map/import runs on accelerator migration tooling
  - HA, observability and backup/DR ship as accelerator baselines
  - release/enablement (image rebuild, per-tenant flags, BigQuery views) is
    scripted rather than hand-run
  - runbook and training material start from templates

Effort drops 214 -> 136 mandays and the calendar drops 14 -> 8 weeks. Effort
does NOT fall uniformly: Zendesk data fidelity, the SAML IdP integration, the
IVR-4 root cause and client UAT time are not things an accelerator compresses,
so those keep close to their baseline estimates and their risk buffers.

Per-phase risk buffer (Risk % -> Contingency mandays) sits on top of the base
estimate; every number stays formula-driven and editable in the sheet.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE = "/Users/yudaadipratama/Archive/id-crm-ticketing/docs/proposals"
OUT = f"{BASE}/PROTON - GCP Pricing, Timeline and PS Utilization (Accelerated 2-Month).xlsx"

RATE = 16500  # USD -> IDR (editable)

# ---- palette (from template) ----
BLUE   = "FF4285F4"; BLUE2 = "FF3C78D8"; YELLOW = "FFFFF2CC"; PINK = "FFF4CCCC"
GREY1  = "FFCCCCCC"; GREY2 = "FFD9D9D9"; GREENDK = "FF274E13"; GREENLT = "FFD9EAD3"
BAR    = "FF9FC5F8"; BARQ = "FFE6E6E6"; WHITE = "FFFFFFFF"; GREYTX = "FF666666"
GS = "Google Sans"; AR = "Arial"
USD_FMT = '[$USD ]#,##0.00'; IDR_FMT = '[$IDR ]#,##0.00'
RP_FMT  = '[$Rp ]#,##0.00'; RP0_FMT = '[$Rp ]#,##0'; PCT = '0%'

thin = Side(style="thin", color="FFB7B7B7")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell(ws, r, c, v=None, *, fill=None, font_name=AR, size=10, bold=False, color=None,
         border=BOX, ha=None, va="center", wrap=False, nf=None):
    cl = ws.cell(r, c)
    if v is not None:
        cl.value = v
    cl.font = Font(name=font_name, size=size, bold=bold, color=color)
    if fill:
        cl.fill = PatternFill("solid", fgColor=fill)
    if border:
        cl.border = border
    cl.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if nf:
        cl.number_format = nf
    return cl


def box(ws, r1, c1, r2, c2, **kw):
    """Style every cell in a rect then merge (keeps full outline on merged range)."""
    v = kw.pop("v", None)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell(ws, r, c, **kw)
    if v is not None:
        ws.cell(r1, c1).value = v
    if (r1, c1) != (r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def fit_landscape(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.2
    ws.page_margins.top = ws.page_margins.bottom = 0.3


wb = openpyxl.Workbook()

# ============================================================ TAB 1: GCP COST
# Unchanged from the baseline workbook: the running platform costs the same
# regardless of which slice of PS scope is being priced. Splitting infra between
# this workbook and the A-G one would double-count it.
ws = wb.active
ws.title = "GCP Cost Estimation"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 5, "C": 30, "D": 52, "E": 15, "F": 17})
ws["A1"] = RATE
ws["A1"].font = Font(name=AR, size=9, color="FF999999")


def gcp_table(top, title, rows, note=None):
    r = top
    if title:
        box(ws, r, 2, r, 6, v=title, fill=GREY2, font_name=AR, bold=True, ha="left")
        r += 1
    hdr = ["No", "GCP Component", "Component Details", "Price (USD)", "Price (IDR)"]
    for i, h in enumerate(hdr):
        cell(ws, r, 2 + i, h, fill=BLUE, font_name=AR, bold=True, color=WHITE, ha="center", wrap=True)
    r += 1
    first = r
    for i, (comp, det, usd) in enumerate(rows, 1):
        cell(ws, r, 2, i, ha="center")
        cell(ws, r, 3, comp, wrap=True, ha="left")
        cell(ws, r, 4, det, wrap=True, ha="left")
        cell(ws, r, 5, usd, nf=USD_FMT, ha="center")
        cell(ws, r, 6, f"=E{r}*A$1", nf=IDR_FMT, ha="center")
        ws.row_dimensions[r].height = 14 * (1 + det.count("\n"))
        r += 1
    last = r - 1
    box(ws, r, 2, r, 4, v="Total Price Estimation Per-Month", fill=YELLOW, bold=True, ha="right")
    cell(ws, r, 5, f"=SUM(E{first}:E{last})", fill=YELLOW, bold=True, nf=USD_FMT, ha="right")
    cell(ws, r, 6, f"=SUM(F{first}:F{last})", fill=YELLOW, bold=True, nf=IDR_FMT, ha="right")
    my = r; r += 1
    box(ws, r, 2, r, 4, v="Total Price Estimation Per-Year", fill=YELLOW, bold=True, ha="right")
    cell(ws, r, 5, f"=E{my}*12", fill=YELLOW, bold=True, nf=USD_FMT, ha="right")
    cell(ws, r, 6, f"=F{my}*12", fill=YELLOW, bold=True, nf=IDR_FMT, ha="right")
    r += 1
    box(ws, r, 2, r, 6, v="Calculator Link:  https://cloud.google.com/products/calculator",
        fill=None, ha="left")
    r += 1
    if note:
        cell(ws, r, 2, note, border=None, font_name=AR, size=9, color="FF999999", ha="left")
        r += 1
    return my, r


recurring = [
    ("Google Kubernetes Engine (Standard) — cluster",
     "Region: Singapore (asia-southeast1)\nRegional cluster management fee ($0.10/cluster/hr)\nMulti-tenant: per-tenant namespaces + node pools", 73.00),
    ("Google Kubernetes Engine (Standard) — node pool",
     "3 × e2-standard-4 (vCPUs: 4, RAM: 16 GiB), on-demand\nRuns Chatwoot (Rails+Sidekiq), agent & backend (FastAPI)", 362.01),
    ("Cloud SQL for PostgreSQL (HA)",
     "Region: Singapore (asia-southeast1)\nEnterprise edition, regional HA\nMachine: 2 vCPU / 8 GiB RAM\nStorage: 100 GB SSD\nChatwoot DB + per-tenant pgvector KB", 300.00),
    ("Memorystore for Redis (HA)",
     "Standard tier (auto-failover replica)\nCapacity: 4 GB\nSidekiq queues / cache", 181.04),
    ("Vertex AI — Gemini 2.5 Flash + Vertex AI Search",
     "Usage-based (indicative)\n~5,000 AI conversations/mo (~8k in + ~1.5k out tokens)\n~5,000 grounded KB queries/mo", 50.75),
    ("Cloud Storage",
     "Region: Singapore (asia-southeast1)\nClass: Standard · 100 GB + modest egress\nAttachments / exports", 8.30),
    ("Networking — Cloud Load Balancing + egress",
     "1 external HTTP(S) forwarding rule + data processing\n~100 GB internet egress/mo", 31.85),
    ("Cloud Logging + Cloud Monitoring",
     "~100 GB log ingest/mo (first 50 GB free)\nMetrics + alerting", 25.00),
    ("Secret Manager + Artifact Registry + Cloud Build + BigQuery",
     "Secrets · container images · CI builds · analytics warehouse (storage + query)", 17.25),
]
onetime = [
    ("Cloud Storage — migration staging",
     "Zendesk export staging bucket (Standard, ~50 GB, short-lived)", 5.00),
    ("Data transfer / egress — migration",
     "One-time import/export bandwidth for tickets, contacts, KB, attachments", 15.00),
    ("Temporary migration compute",
     "Short-lived job for bulk import + KB ingest into Vertex AI Search / pgvector", 20.00),
]

my1, nxt = gcp_table(2, "PROTON CRM — Monthly Recurring (one production tenant)", recurring)
_, nxt2 = gcp_table(nxt + 1, "One-time — Migration Period *", onetime,
                    note="*one time only during migration process")
box(ws, nxt2 + 1, 2, nxt2 + 1, 6,
    v=f"IDR shown at USD→IDR = cell A1 ({RATE:,}, editable). Indicative GCP list price (Jul 2026, asia-southeast1 / "
      f"Singapore — nearest full-service region; Malaysia region not yet GA). Excludes negotiated discounts, taxes & "
      f"Google program funding. 1-yr committed-use discount lowers nodes ~$233 → ~$920/mo total. Infrastructure is "
      f"unchanged from the baseline workbook — it carries the whole platform, so it is not split between this scope "
      f"and the Packages A-G scope.",
    fill=None, border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws.row_dimensions[nxt2 + 1].height = 52
fit_landscape(ws)

# ============================================================ TAB 2: TIMELINE
ws2 = wb.create_sheet("Timeline")
ws2.sheet_view.showGridLines = False
NW = 8  # 2 months
widths(ws2, {"B": 4, "C": 5, "D": 60})
for i in range(NW):
    ws2.column_dimensions[get_column_letter(5 + i)].width = 5.4

box(ws2, 2, 2, 2, 3, v="Phases", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")
cell(ws2, 2, 4, "Task", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")
for i in range(NW):
    cell(ws2, 2, 5 + i, f"W{i+1}", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")

# phases: (name, phase_no, [(subno, task, wk_start, wk_end)])
phases = [
    ("Assessment & Planning", 1, [
        ("1.1", "Kick-off workshop; confirm scope against the BRD + Process-Flow workbook", 1, 1),
        ("1.2", "Validate the 5-channel complaint SOP and the escalation matrix", 1, 1),
        ("1.3", "Solution architecture, GCP landing-zone design & migration plan", 1, 1),
    ]),
    ("GCP Foundation & Provisioning", 2, [
        ("2.1", "Provision GCP project, VPC, IAM & landing zone from accelerator IaC", 1, 2),
        ("2.2", "GKE Standard cluster + node pools; Cloud SQL (HA), Memorystore, BigQuery, Artifact Registry", 2, 2),
        ("2.3", "CI/CD (Cloud Build), Secret Manager, monitoring baseline & environment smoke test", 2, 2),
    ]),
    ("Core Platform Cutover", 3, [
        ("3.1", "Deploy Chatwoot + agent + backend on GKE via accelerator Helm charts", 2, 3),
        ("3.2", "Migrate Zendesk tickets, contacts, history & attachments", 3, 4),
        ("3.3", "Migrate the knowledge base to pgvector + Vertex AI Search", 3, 4),
        ("3.4", "Wire channels: WhatsApp (Twilio), Website widget, API, Email", 3, 4),
        ("3.5", "Enable the AI layer & SOP lifecycle flows; reconcile migrated records", 4, 4),
    ]),
    ("Gap Closure", 4, [
        ("4.1", "Email escalation — auto-ack once per thread + separate internal case", 4, 5),
        ("4.2", "Customer 360 card and the DMS/TSP integration surface", 4, 5),
        ("4.3", "Case management — hierarchical categories, per-case dashboard, RBAC", 5, 5),
        ("4.4", "Reporting & BI; agent management; bulk FAQ upload, multi-language & multimodal", 5, 5),
    ]),
    ("Hardening, HA & UAT", 5, [
        ("5.1", "HA / autoscaling, resource tuning; monitoring, alerting & logging", 5, 6),
        ("5.2", "Backups & disaster-recovery runbook, with a restore test", 6, 6),
        ("5.3", "Security review, RBAC verification & secret rotation", 6, 6),
        ("5.4", "Live channel verification (WhatsApp, phone, email, web)", 6, 7),
        ("5.5", "UAT cycles, defect triage & fixes; data reconciliation & sign-off", 6, 7),
    ]),
    ("Cutover, Handover & Training", 6, [
        ("6.1", "Cutover plan, go-live & hypercare window", 7, 7),
        ("6.2", "Zendesk decommission & licence termination", 8, 8),
        ("6.3", "Ops runbook; admin & agent training; knowledge transfer & closure", 7, 8),
    ]),
]
phase_fills = [GREY1, GREY2]
r = 3
for pi, (pname, pno, tasks) in enumerate(phases):
    box(ws2, r, 2, r, 4, v=pname, fill=phase_fills[pi % 2], font_name=GS, bold=True, color=GREYTX, ha="center")
    for i in range(NW):
        cell(ws2, r, 5 + i, fill=phase_fills[pi % 2])
    r += 1
    start_task_row = r
    for (subno, task, ws_, we_) in tasks:
        cell(ws2, r, 3, subno, font_name=GS, color=GREYTX, ha="center")
        cell(ws2, r, 4, task, font_name=GS, color=GREYTX, ha="left", wrap=True)
        for i in range(NW):
            wk = i + 1
            if ws_ <= wk <= we_:
                cell(ws2, r, 5 + i, fill=BAR)
            else:
                cell(ws2, r, 5 + i)
        ws2.row_dimensions[r].height = 26
        r += 1
    box(ws2, start_task_row, 2, r - 1, 2, v=f"{pno}.0", font_name=GS, bold=True, color=GREYTX, ha="center")

r += 1
cell(ws2, r, 2, "Accelerated delivery — 8 weeks (2 months), 36 working days of duration, against 16 weeks on the "
                "standard-build plan. The compression comes from the accelerator: the landing zone and GKE/Helm "
                "deployment are parameterised rather than authored, migration runs on existing tooling, HA/observability/DR "
                "ship as baselines, and the runbook and training material start from templates. Phases overlap — Gap "
                "Closure begins while the platform cutover is still settling. Assumes PROTON supplies the Zendesk export, "
                "DMS/TSP API access and UAT participants on schedule.",
     border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
ws2.row_dimensions[r].height = 44
fit_landscape(ws2)

# ============================================================ TAB 3: PS UTILIZATION
ws3 = wb.create_sheet("PS Utilization")
ws3.sheet_view.showGridLines = False
widths(ws3, {"A": 9, "B": 26, "C": 58, "D": 20, "E": 11, "F": 18, "G": 2,
             "H": 12, "I": 15, "J": 15, "K": 15, "L": 15, "M": 15, "N": 15, "O": 16,
             "P": 2, "Q": 20, "R": 18, "S": 15})
ws3["A1"] = RATE
ws3["A1"].font = Font(name=AR, size=9, color="FF999999")

# ---- rate card (Q:S) ----
rate_rows = [
    ("Consultant", 350000),
    ("Senior Consultant", 606250),
    ("Lead Consultant", 700000),
    ("Principal", 825000),
    ("Expert / Director", 1275000),
]
cell(ws3, 2, 17, "Role / Seniority", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="left")
cell(ws3, 2, 18, "Sales price / hour (IDR)", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="center", wrap=True)
cell(ws3, 2, 19, "Per day (× 8)", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="center", wrap=True)
rr = 3
rate_cell = {}
for name, rate in rate_rows:
    cell(ws3, rr, 17, name, fill=GREENDK, font_name=GS, size=9, bold=True, color=WHITE, border=None, ha="left")
    cell(ws3, rr, 18, rate, fill=GREENLT, font_name=GS, size=9, border=None, ha="center", nf=RP0_FMT)
    cell(ws3, rr, 19, f"=R{rr}*8", fill=GREENLT, font_name=GS, size=9, border=None, ha="center", nf=RP0_FMT)
    rate_cell[name] = f"$R${rr}"
    rr += 1
cell(ws3, rr, 17, "Margin", fill=GREENDK, font_name=GS, size=9, bold=True, color=WHITE, border=None, ha="left")
cell(ws3, rr, 18, 0.15, fill=GREENLT, font_name=AR, bold=True, border=None, ha="right", nf=PCT)
MARGIN = f"$R${rr}"

# ---- role columns (I..N) : (header seniority, header function, rate_cell) ----
# Solution Architect is priced as Lead Consultant here, per PROTON's reference sheet
# (the baseline workbook had it at Principal).
roles = [
    ("Lead Consultant", "Project Manager", rate_cell["Lead Consultant"]),
    ("Lead Consultant", "Solution Architect / Tech Lead", rate_cell["Lead Consultant"]),
    ("Senior Consultant", "Backend / AI Engineer", rate_cell["Senior Consultant"]),
    ("Consultant", "Frontend Engineer", rate_cell["Consultant"]),
    ("Senior Consultant", "Data / BI Engineer", rate_cell["Senior Consultant"]),
    ("Consultant", "QA Engineer", rate_cell["Consultant"]),
]

# ---- phases: (milestone, activity task-list, resources, duration days, [md per role]) ----
# Mandays and durations are PROTON's reference figures (120 MD over 36 working days).
ps_phases = [
    ("Phase 1 : Assessment & Planning",
     "•  Kick-off workshop; confirm scope against the BRD + Process-Flow workbook\n"
     "•  Validate the 5-channel complaint SOP and the escalation matrix\n"
     "•  Solution architecture & GCP landing-zone design (GKE Standard, VPC, IAM)\n"
     "•  Migration plan and Zendesk data-mapping approach\n"
     "•  Delivery plan, RAID log and acceptance criteria agreed",
     "Project Manager\nSolution Architect\nBackend/AI Engineer", 3, [2, 3, 3, 0, 0, 0]),
    ("Phase 2 : GCP Foundation & Provisioning",
     "•  Provision GCP project, VPC, IAM and landing zone from accelerator IaC\n"
     "•  Stand up GKE Standard cluster + node pools\n"
     "•  Provision Cloud SQL (HA), Memorystore, BigQuery and Artifact Registry\n"
     "•  Configure CI/CD (Cloud Build), Secret Manager and the monitoring baseline\n"
     "•  Environment smoke test and access handover",
     "Solution Architect\nDevOps / Backend\nData/BI Engineer\nProject Manager", 3, [2, 3, 3, 1, 2, 1]),
    ("Phase 3 : Core Platform Cutover",
     "•  Deploy Chatwoot + agent + backend on GKE via accelerator Helm charts\n"
     "•  Migrate Zendesk tickets, contacts, history and attachments\n"
     "•  Migrate the knowledge base to pgvector + Vertex AI Search\n"
     "•  Wire channels: WhatsApp (Twilio), Website widget, API, Email\n"
     "•  Enable the AI layer and SOP lifecycle flows (disclaimer, idle-close, YES/NO, rating, auto-classify)\n"
     "•  Reconcile migrated records and confirm counts",
     "Project Manager\nBackend/AI Engineer\nFrontend Engineer\nData/BI Engineer\nQA", 10, [5, 1, 10, 10, 2, 2]),
    ("Phase 4 : Gap Closure",
     "•  Email escalation — auto-ack once per thread + separate internal case\n"
     "•  Customer 360 card and the DMS/TSP integration surface\n"
     "•  Case management — hierarchical categories, per-case dashboard, RBAC\n"
     "•  Reporting & BI — BigQuery views, role-based dashboards, NPS/CRR\n"
     "•  Agent management — channel priority, auto-busy, status labels\n"
     "•  Bulk FAQ/PDF upload; multi-language and multimodal handling",
     "Backend/AI Engineer\nFrontend Engineer\nData/BI Engineer\nSolution Architect\nQA", 5, [5, 3, 5, 5, 2, 2]),
    ("Phase 5 : Hardening, HA & UAT",
     "•  HA / autoscaling and resource tuning\n"
     "•  Monitoring, alerting and centralised logging\n"
     "•  Backups and disaster-recovery runbook, with a restore test\n"
     "•  Security review, RBAC verification and secret rotation\n"
     "•  Live channel verification (WhatsApp, phone, email, web)\n"
     "•  UAT cycles, defect triage and fixes; data reconciliation and sign-off",
     "DevOps / Backend\nQA\nBackend/AI Engineer\nData/BI Engineer\nProject Manager", 10, [5, 2, 10, 5, 5, 5]),
    ("Phase 6 : Cutover, Handover & Training",
     "•  Cutover plan, go-live and the hypercare window\n"
     "•  Zendesk decommission and licence termination\n"
     "•  Ops runbook — deploy, backup/restore, monitoring, tenant provisioning\n"
     "•  Admin and agent training sessions\n"
     "•  Knowledge transfer and project closure report",
     "Project Manager\nSolution Architect\nBackend/AI Engineer\nData/BI Engineer", 5, [4, 5, 5, 0, 2, 0]),
]

# title + headers
box(ws3, 3, 2, 3, 6, v="Professional Services — PROTON CRM (self-managed on Google Cloud) · accelerated delivery, 2 months",
    fill=PINK, font_name=GS, bold=True, wrap=True, ha="center")
for c, h in [(2, "Milestone"), (3, "Activity"), (4, "Resources"), (5, "Duration (Week)"), (6, "Service Price")]:
    cell(ws3, 4, c, h, fill=PINK, font_name=GS, bold=True, ha="center", wrap=True)
cell(ws3, 4, 8, "Duration (Days)", fill=YELLOW, font_name=GS, bold=True, ha="center", wrap=True)
for i, (sen, fn, _) in enumerate(roles):
    cell(ws3, 3, 9 + i, sen, fill=YELLOW, font_name=GS, size=9, bold=True, ha="center", wrap=True)
    cell(ws3, 4, 9 + i, fn, fill=YELLOW, font_name=GS, size=9, bold=True, ha="center", wrap=True)
cell(ws3, 3, 15, "Total", fill=YELLOW, font_name=GS, bold=True, ha="center")
cell(ws3, 4, 15, "(mandays)", fill=YELLOW, font_name=GS, size=9, bold=True, ha="center")

first = 5
r = first
for (mil, act, res, days, mds) in ps_phases:
    cell(ws3, r, 2, mil, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 3, act, font_name=GS, size=9, wrap=True, ha="left")
    cell(ws3, r, 4, res, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 5, f"=ROUNDUP(H{r}/5,0)", font_name=GS, ha="center", nf='#,##0')
    cell(ws3, r, 8, days, font_name=GS, ha="center")
    for i, md in enumerate(mds):
        cell(ws3, r, 9 + i, md, font_name=GS, ha="center")
    cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=GS, ha="center")
    ws3.row_dimensions[r].height = 84
    r += 1
last = r - 1

# Service Price = one merged cell (F5:F10) = total incl. margin (row last+3), rounded
box(ws3, first, 6, last, 6, v=f"=ROUNDUP(O{last+3},-2)", font_name=GS, bold=True, ha="center", nf=RP0_FMT)

# ---- totals block ----
cell(ws3, r, 8, "Total Mandays", fill=YELLOW, font_name=AR, bold=True, ha="left")
for i in range(len(roles)):
    cl = get_column_letter(9 + i)
    cell(ws3, r, 9 + i, f"=SUM({cl}{first}:{cl}{last})", font_name=AR, ha="center")
cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=AR, bold=True, ha="center")
row_md = r; r += 1

cell(ws3, r, 8, "Total Mandays Price", fill=YELLOW, font_name=AR, bold=True, ha="left")
for i, (_, _, ratec) in enumerate(roles):
    cl = get_column_letter(9 + i)
    cell(ws3, r, 9 + i, f"={cl}{row_md}*8*{ratec}", font_name=AR, ha="center", nf=RP0_FMT)
cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=AR, bold=True, ha="center", nf=RP0_FMT)
row_price = r; r += 1

cell(ws3, r, 8, "Total Price Including Margin", fill=YELLOW, font_name=AR, bold=True, ha="left", wrap=True)
box(ws3, r, 9, r, 14, v=None)
cell(ws3, r, 15, f"=O{row_price}*(1+{MARGIN})", font_name=AR, bold=True, ha="center", nf=RP0_FMT)
row_total = r; r += 2

cell(ws3, r, 2, "Accelerated delivery: 120 mandays over 36 working days (~2 months), against 252 mandays over 16 weeks "
                "on the standard-build plan. The compression comes from the accelerator — the landing zone and GKE/Helm "
                "deployment are parameterised rather than authored, migration runs on existing tooling, HA/observability/DR "
                "ship as baselines, release is scripted, and the runbook and training material start from templates. "
                "Assumes PROTON supplies the Zendesk export, DMS/TSP API access and UAT participants on schedule; slippage "
                "on any of those moves the go-live date rather than being absorbed. Mandays are an indicative delivery "
                "estimate; day-rates and margin are editable in the rate card (Q:S). Note the Solution Architect is priced "
                "at the Lead Consultant rate. May be offset by Google program funding.",
     border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=15)
ws3.row_dimensions[r].height = 70
fit_landscape(ws3)

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
