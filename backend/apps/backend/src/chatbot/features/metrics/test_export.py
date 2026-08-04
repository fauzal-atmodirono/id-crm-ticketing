import asyncio
import csv
import io
from dataclasses import fields
from datetime import date

from openpyxl import load_workbook  # type: ignore[import-untyped]

from chatbot.features.metrics.export import render_csv, render_pdf, render_xlsx
from chatbot.features.metrics.query_port import (
    CsatRow,
    DashboardMetrics,
    DealerEscalationMetrics,
    DealerEscalationRow,
    DealerSlowCaseRow,
    MockMetricsQuery,
    VolumeByTypeDivisionMetrics,
    VolumeByTypeDivisionRow,
    VolumeRow,
)


async def _metrics() -> DashboardMetrics:
    return await MockMetricsQuery().fetch_dashboard()


def _sync_metrics() -> DashboardMetrics:
    return asyncio.run(_metrics())


def test_render_xlsx_is_a_zip_workbook() -> None:
    data = render_xlsx(_sync_metrics())
    assert data[:2] == b"PK"  # xlsx is a zip container
    assert len(data) > 500


def test_render_xlsx_has_a_sheet_per_block() -> None:
    wb = load_workbook(io.BytesIO(render_xlsx(_sync_metrics())))
    # DashboardMetrics.scopes (Task 2 / Package E) is a property, not a
    # dataclass field -- see its docstring in query_port.py -- so
    # fields(DashboardMetrics) never includes it and there's no sheet for
    # it, without _blocks() needing to filter anything out.
    for block in [f.name for f in fields(DashboardMetrics)]:
        assert block in wb.sheetnames
    assert "scopes" not in wb.sheetnames


def test_render_pdf_has_pdf_header() -> None:
    data = render_pdf(_sync_metrics())
    assert data[:5] == b"%PDF-"
    assert len(data) > 500


def test_render_csv_produces_one_section_per_block():
    metrics = DealerEscalationMetrics(
        by_dealer=[DealerEscalationRow("Dealer KL", 12, 3.5, 3.0, 6.0)],
        slowest_cases=[DealerSlowCaseRow("CONV042", "Dealer KL", 12.0)],
    )
    content = render_csv(metrics).decode("utf-8")
    reader = list(csv.reader(io.StringIO(content)))
    assert ["by_dealer"] in reader
    assert ["dealer", "cases_escalated", "avg_turnaround_days", "p50_turnaround_days", "p90_turnaround_days"] in reader
    assert ["Dealer KL", "12", "3.5", "3.0", "6.0"] in reader
    assert ["slowest_cases"] in reader


def test_render_csv_empty_block_has_no_data_marker():
    metrics = DealerEscalationMetrics(by_dealer=[], slowest_cases=[])
    content = render_csv(metrics).decode("utf-8")
    assert "(no data)" in content


def test_render_csv_drops_a_period_only_column_deterministically() -> None:
    """VolumeRow.bucket is only ever populated by a period-scoped query
    (see its field metadata in query_port.py) -- it must be dropped from
    an unfiltered export regardless of the data, not because it happens
    to be None on every row this time."""
    metrics = DashboardMetrics(
        volume=[
            VolumeRow(month="2026-05", channel="web", volume=120),
            VolumeRow(month="2026-06", channel="whatsapp", volume=95),
        ],
        resolution=[],
        csat=[],
        nps=[],
        speed=[],
        fallback=[],
        bounce=[],
        quality=[],
    )
    content = render_csv(metrics).decode("utf-8")
    reader = list(csv.reader(io.StringIO(content)))
    assert ["month", "channel", "volume"] in reader
    for row in reader:
        assert "bucket" not in row


def test_render_csv_keeps_a_nullable_business_column_even_when_all_none() -> None:
    """Task 2 review, round 3: the distinction that matters is whether a
    field is *structurally* period-only (VolumeRow.bucket), not whether it
    merely *happens* to be None on every row of a particular export.
    CsatRow.avg_score/satisfied_rate are SAFE_DIVIDE columns, legitimately
    null when their denominator is zero (see query_port.py's module
    docstring) -- a quiet week or small tenant can plausibly have every
    row null for one of these, and the column must still appear, or
    anything parsing the emailed report by header/position breaks on a
    schema that silently shifts week to week."""
    metrics = DashboardMetrics(
        volume=[],
        resolution=[],
        csat=[
            CsatRow(channel="web", respondents=0, avg_score=None, satisfied_rate=None),
            CsatRow(channel="whatsapp", respondents=0, avg_score=None, satisfied_rate=None),
        ],
        nps=[],
        speed=[],
        fallback=[],
        bounce=[],
        quality=[],
    )
    content = render_csv(metrics).decode("utf-8")
    reader = list(csv.reader(io.StringIO(content)))
    assert ["channel", "respondents", "avg_score", "satisfied_rate"] in reader


def test_export_keeps_month_start_but_drops_the_period_only_bucket() -> None:
    """Package E final fix, finding M2 -- the decision, pinned.

    `VolumeByTypeDivisionRow` gained two fields. They are exported
    differently, on purpose:

    - `bucket` is `period_only`: no export route ever supplies a period
      (`export_router.py` calls every `fetch_*` with no arguments), so it
      could never be anything but blank here. Dropped, same as
      `VolumeRow.bucket`.
    - `month_start` is NOT `period_only`: it is selected by the month-grain
      view `v_volume_by_type_division` and is therefore genuinely populated
      on exactly the unfiltered path exports use. Dropping it would violate
      `_exportable_field_names`'s stated contract -- it only ever removes a
      column that is structurally blank, never real data that merely looks
      redundant.

    So `/metrics/volume-by-type/export`'s CSV header gains `month_start`
    (and only that) with this branch. Logged in the spec's deploy steps.
    """
    metrics = VolumeByTypeDivisionMetrics(
        volume=[
            VolumeByTypeDivisionRow(
                month="2026-06",
                channel="WhatsApp",
                case_type="Inquiry",
                division="Sales",
                volume=682,
                month_start=date(2026, 6, 1),
            )
        ]
    )
    reader = list(csv.reader(io.StringIO(render_csv(metrics).decode("utf-8"))))
    assert [
        "month",
        "channel",
        "case_type",
        "division",
        "volume",
        "month_start",
    ] in reader
    for row in reader:
        assert "bucket" not in row
