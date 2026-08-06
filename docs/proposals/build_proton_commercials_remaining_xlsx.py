#!/usr/bin/env python3
"""Build the PROTON commercials workbook for the scope OUTSIDE Packages A-G.

Same 3-tab presales format as `build_proton_commercials_xlsx.py`
(GCP Cost Estimation · Timeline · PS Utilization), but zero-based on the work
that is *not* covered by the 2026-08-04 package queue (A-G), which is already
in delivery and priced separately.

What that leaves:
  1. Platform delivery that has never started — GCP landing zone, GKE
     re-platform off Docker Compose, Zendesk + KB migration, channel cutover,
     hardening/HA/UAT, go-live, decommission, handover.
  2. Product items with a spec and/or plan in docs/superpowers/ but no code:
     native SAML SSO, RBAC phase 3 (native conversation visibility), the
     IVR-4 voice-language defect, custom agent status labels, plus the
     release/enablement and live-verification backlog.

Adds a per-phase risk buffer (Risk % -> Contingency mandays) on top of the
base estimate; every number stays formula-driven and editable in the sheet.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE = "/Users/yudaadipratama/Archive/id-crm-ticketing/docs/proposals"
OUT = f"{BASE}/PROTON - GCP Pricing, Timeline and PS Utilization (Remaining Scope).xlsx"

RATE = 16500  # USD -> IDR (editable)

# ---- palette (from template) ----
BLUE   = "FF4285F4"; BLUE2 = "FF3C78D8"; YELLOW = "FFFFF2CC"; PINK = "FFF4CCCC"
GREY1  = "FFCCCCCC"; GREY2 = "FFD9D9D9"; GREENDK = "FF274E13"; GREENLT = "FFD9EAD3"
BAR    = "FF9FC5F8"; BARQ = "FFE6E6E6"; WHITE = "FFFFFFFF"; GREYTX = "FF666666"
GS = "Google Sans"; AR = "Arial"
USD_FMT = '[$USD ]#,##0.00'; IDR_FMT = '[$IDR ]#,##0.00'
RP_FMT  = '[$Rp ]#,##0.00'; RP0_FMT = '[$Rp ]#,##0'; PCT = '0%'

thin = Side(style="thin", color="FFB7B7B7")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)


def cell(ws, r, c, v=None, *, fill=None, font_name=AR, size=10, bold=False, color=None,
         border=BOX, ha=None, va="center", wrap=False, nf=None):
    cl = ws.cell(r, c)
    if v is not None:
        cl.value = v
    cl.font = Font(name=font_name, size=size, bold=bold, color=color)
    if fill:
        cl.fill = PatternFill("solid", fgColor=fill)
    if border:
        cl.border = border
    cl.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if nf:
        cl.number_format = nf
    return cl


def box(ws, r1, c1, r2, c2, **kw):
    """Style every cell in a rect then merge (keeps full outline on merged range)."""
    v = kw.pop("v", None)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell(ws, r, c, **kw)
    if v is not None:
        ws.cell(r1, c1).value = v
    if (r1, c1) != (r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def fit_landscape(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.2
    ws.page_margins.top = ws.page_margins.bottom = 0.3


wb = openpyxl.Workbook()

# ============================================================ TAB 1: GCP COST
# Unchanged from the baseline workbook: the running platform costs the same
# regardless of which slice of PS scope is being priced. Splitting infra between
# this workbook and the A-G one would double-count it.
ws = wb.active
ws.title = "GCP Cost Estimation"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 5, "C": 30, "D": 52, "E": 15, "F": 17})
ws["A1"] = RATE
ws["A1"].font = Font(name=AR, size=9, color="FF999999")


def gcp_table(top, title, rows, note=None):
    r = top
    if title:
        box(ws, r, 2, r, 6, v=title, fill=GREY2, font_name=AR, bold=True, ha="left")
        r += 1
    hdr = ["No", "GCP Component", "Component Details", "Price (USD)", "Price (IDR)"]
    for i, h in enumerate(hdr):
        cell(ws, r, 2 + i, h, fill=BLUE, font_name=AR, bold=True, color=WHITE, ha="center", wrap=True)
    r += 1
    first = r
    for i, (comp, det, usd) in enumerate(rows, 1):
        cell(ws, r, 2, i, ha="center")
        cell(ws, r, 3, comp, wrap=True, ha="left")
        cell(ws, r, 4, det, wrap=True, ha="left")
        cell(ws, r, 5, usd, nf=USD_FMT, ha="center")
        cell(ws, r, 6, f"=E{r}*A$1", nf=IDR_FMT, ha="center")
        ws.row_dimensions[r].height = 14 * (1 + det.count("\n"))
        r += 1
    last = r - 1
    box(ws, r, 2, r, 4, v="Total Price Estimation Per-Month", fill=YELLOW, bold=True, ha="right")
    cell(ws, r, 5, f"=SUM(E{first}:E{last})", fill=YELLOW, bold=True, nf=USD_FMT, ha="right")
    cell(ws, r, 6, f"=SUM(F{first}:F{last})", fill=YELLOW, bold=True, nf=IDR_FMT, ha="right")
    my = r; r += 1
    box(ws, r, 2, r, 4, v="Total Price Estimation Per-Year", fill=YELLOW, bold=True, ha="right")
    cell(ws, r, 5, f"=E{my}*12", fill=YELLOW, bold=True, nf=USD_FMT, ha="right")
    cell(ws, r, 6, f"=F{my}*12", fill=YELLOW, bold=True, nf=IDR_FMT, ha="right")
    r += 1
    box(ws, r, 2, r, 6, v="Calculator Link:  https://cloud.google.com/products/calculator",
        fill=None, ha="left")
    r += 1
    if note:
        cell(ws, r, 2, note, border=None, font_name=AR, size=9, color="FF999999", ha="left")
        r += 1
    return my, r


recurring = [
    ("Google Kubernetes Engine (Standard) — cluster",
     "Region: Singapore (asia-southeast1)\nRegional cluster management fee ($0.10/cluster/hr)\nMulti-tenant: per-tenant namespaces + node pools", 73.00),
    ("Google Kubernetes Engine (Standard) — node pool",
     "3 × e2-standard-4 (vCPUs: 4, RAM: 16 GiB), on-demand\nRuns Chatwoot (Rails+Sidekiq), agent & backend (FastAPI)", 362.01),
    ("Cloud SQL for PostgreSQL (HA)",
     "Region: Singapore (asia-southeast1)\nEnterprise edition, regional HA\nMachine: 2 vCPU / 8 GiB RAM\nStorage: 100 GB SSD\nChatwoot DB + per-tenant pgvector KB", 300.00),
    ("Memorystore for Redis (HA)",
     "Standard tier (auto-failover replica)\nCapacity: 4 GB\nSidekiq queues / cache", 181.04),
    ("Vertex AI — Gemini 2.5 Flash + Vertex AI Search",
     "Usage-based (indicative)\n~5,000 AI conversations/mo (~8k in + ~1.5k out tokens)\n~5,000 grounded KB queries/mo", 50.75),
    ("Cloud Storage",
     "Region: Singapore (asia-southeast1)\nClass: Standard · 100 GB + modest egress\nAttachments / exports", 8.30),
    ("Networking — Cloud Load Balancing + egress",
     "1 external HTTP(S) forwarding rule + data processing\n~100 GB internet egress/mo", 31.85),
    ("Cloud Logging + Cloud Monitoring",
     "~100 GB log ingest/mo (first 50 GB free)\nMetrics + alerting", 25.00),
    ("Secret Manager + Artifact Registry + Cloud Build + BigQuery",
     "Secrets · container images · CI builds · analytics warehouse (storage + query)", 17.25),
]
onetime = [
    ("Cloud Storage — migration staging",
     "Zendesk export staging bucket (Standard, ~50 GB, short-lived)", 5.00),
    ("Data transfer / egress — migration",
     "One-time import/export bandwidth for tickets, contacts, KB, attachments", 15.00),
    ("Temporary migration compute",
     "Short-lived job for bulk import + KB ingest into Vertex AI Search / pgvector", 20.00),
]

my1, nxt = gcp_table(2, "PROTON CRM — Monthly Recurring (one production tenant)", recurring)
_, nxt2 = gcp_table(nxt + 1, "One-time — Migration Period *", onetime,
                    note="*one time only during migration process")
box(ws, nxt2 + 1, 2, nxt2 + 1, 6,
    v=f"IDR shown at USD→IDR = cell A1 ({RATE:,}, editable). Indicative GCP list price (Jul 2026, asia-southeast1 / "
      f"Singapore — nearest full-service region; Malaysia region not yet GA). Excludes negotiated discounts, taxes & "
      f"Google program funding. 1-yr committed-use discount lowers nodes ~$233 → ~$920/mo total. Infrastructure is "
      f"unchanged from the baseline workbook — it carries the whole platform, so it is not split between this scope "
      f"and the Packages A-G scope.",
    fill=None, border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws.row_dimensions[nxt2 + 1].height = 52
fit_landscape(ws)

# ============================================================ TAB 2: TIMELINE
ws2 = wb.create_sheet("Timeline")
ws2.sheet_view.showGridLines = False
NW = 14
widths(ws2, {"B": 4, "C": 5, "D": 46})
for i in range(NW):
    ws2.column_dimensions[get_column_letter(5 + i)].width = 3.8

box(ws2, 2, 2, 2, 3, v="Phases", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")
cell(ws2, 2, 4, "Task", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")
for i in range(NW):
    cell(ws2, 2, 5 + i, f"W{i+1}", fill=BLUE2, font_name=GS, bold=True, color=WHITE, ha="center")

# phases: (name, phase_no, [(subno, task, wk_start, wk_end)])
phases = [
    ("GCP Foundation & Provisioning", 1, [
        ("1.1", "Provision GCP project, VPC, IAM, GKE Standard cluster + node pools", 1, 2),
        ("1.2", "Cloud SQL, Memorystore, BigQuery, Artifact Registry; CI/CD, Secret Manager, monitoring baseline", 2, 3),
        ("1.3", "Re-platform Chatwoot + agent + backend from Docker Compose to GKE manifests", 2, 3),
    ]),
    ("Data Migration & Channel Cutover", 2, [
        ("2.1", "Zendesk extract, field mapping & bulk import (tickets, contacts, attachments)", 4, 5),
        ("2.2", "KB migration to pgvector + Vertex AI Search corpus", 4, 5),
        ("2.3", "Re-wire WhatsApp (Twilio), Website, API & Email inboxes on the new cluster", 5, 6),
        ("2.4", "Parallel-run (Zendesk + new CRM) & data reconciliation", 6, 7),
    ]),
    ("Platform Features outside Packages A–G", 3, [
        ("3.1", "Native SAML SSO — IdP metadata, ACS endpoint, JIT provisioning, Security settings UI", 3, 6),
        ("3.2", "RBAC phase 3 — native conversation-visibility scoping", 5, 8),
        ("3.3", "IVR-4 — voice-language reliability diagnosis & fix", 6, 7),
        ("3.4", "Custom agent status labels (Follow-up / Lunch / Break) + routing integration", 7, 8),
        ("3.5", "Release & enablement — Chatwoot image rebuild (patches 0034–0041), RSA DB provisioning, BigQuery view rollout, per-tenant flags", 8, 10),
    ]),
    ("Hardening, HA & UAT", 4, [
        ("4.1", "HA / autoscaling, resource tuning, monitoring & alerting", 9, 10),
        ("4.2", "Backups & disaster-recovery runbook", 10, 11),
        ("4.3", "Live channel verification (real WhatsApp, phone call, browser smoke)", 10, 11),
        ("4.4", "UAT, data reconciliation & sign-off", 11, 12),
    ]),
    ("Cutover, Decommission & Handover", 5, [
        ("5.1", "Go-live cutover & hypercare", 13, 13),
        ("5.2", "Zendesk decommission", 13, 14),
        ("5.3", "Ops runbook + admin/agent training & knowledge transfer", 13, 14),
    ]),
]
phase_fills = [GREY1, GREY2]
r = 3
for pi, (pname, pno, tasks) in enumerate(phases):
    box(ws2, r, 2, r, 4, v=pname, fill=phase_fills[pi % 2], font_name=GS, bold=True, color=GREYTX, ha="center")
    for i in range(NW):
        cell(ws2, r, 5 + i, fill=phase_fills[pi % 2])
    r += 1
    start_task_row = r
    for (subno, task, ws_, we_) in tasks:
        cell(ws2, r, 3, subno, font_name=GS, color=GREYTX, ha="center")
        cell(ws2, r, 4, task, font_name=GS, color=GREYTX, ha="left", wrap=True)
        for i in range(NW):
            wk = i + 1
            if ws_ <= wk <= we_:
                cell(ws2, r, 5 + i, fill=BAR)
            else:
                cell(ws2, r, 5 + i)
        ws2.row_dimensions[r].height = 26
        r += 1
    box(ws2, start_task_row, 2, r - 1, 2, v=f"{pno}.0", font_name=GS, bold=True, color=GREYTX, ha="center")

# memo band — packages A-G run in parallel but are priced in their own workbook
box(ws2, r, 2, r, 4, v="Memo — priced separately", fill=GREY1, font_name=GS, bold=True, color=GREYTX, ha="center")
for i in range(NW):
    cell(ws2, r, 5 + i, fill=GREY1)
r += 1
cell(ws2, r, 3, "A–G", font_name=GS, color=GREYTX, ha="center")
cell(ws2, r, 4, "Packages A–G (email/social, Contacts 360 merge, telephony, demo data & case detail, "
                "reporting deck parity, DMS/TSP, escalation-policy engine) — already queued, own estimate",
     font_name=GS, color=GREYTX, ha="left", wrap=True)
for i in range(NW):
    cell(ws2, r, 5 + i, fill=BARQ)
ws2.row_dimensions[r].height = 26
box(ws2, r, 2, r, 2, v="—", font_name=GS, bold=True, color=GREYTX, ha="center")
r += 2
cell(ws2, r, 2, "Timeline covers the scope outside Packages A–G only. Phase durations overlap — the sum of phase "
                "weeks exceeds the 14-week calendar because Phase 3 runs alongside Phases 1–2.",
     border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
fit_landscape(ws2)

# ============================================================ TAB 3: PS UTILIZATION
ws3 = wb.create_sheet("PS Utilization")
ws3.sheet_view.showGridLines = False
widths(ws3, {"A": 9, "B": 26, "C": 52, "D": 20, "E": 11, "F": 18, "G": 2,
             "H": 12, "I": 15, "J": 15, "K": 15, "L": 15, "M": 15, "N": 15, "O": 16,
             "P": 9, "Q": 14, "R": 16, "S": 2, "T": 20, "U": 18, "V": 15})
ws3["A1"] = RATE
ws3["A1"].font = Font(name=AR, size=9, color="FF999999")

# ---- rate card (T:V) ----
rate_rows = [
    ("Consultant", 350000),
    ("Senior Consultant", 606250),
    ("Lead Consultant", 700000),
    ("Principal", 825000),
    ("Expert / Director", 1275000),
]
cell(ws3, 2, 20, "Role / Seniority", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="left")
cell(ws3, 2, 21, "Sales price / hour (IDR)", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="center", wrap=True)
cell(ws3, 2, 22, "Per day (× 8)", fill=WHITE, font_name=GS, size=9, bold=True, border=None, ha="center", wrap=True)
rr = 3
rate_cell = {}
for name, rate in rate_rows:
    cell(ws3, rr, 20, name, fill=GREENDK, font_name=GS, size=9, bold=True, color=WHITE, border=None, ha="left")
    cell(ws3, rr, 21, rate, fill=GREENLT, font_name=GS, size=9, border=None, ha="center", nf=RP0_FMT)
    cell(ws3, rr, 22, f"=U{rr}*8", fill=GREENLT, font_name=GS, size=9, border=None, ha="center", nf=RP0_FMT)
    rate_cell[name] = f"$U${rr}"
    rr += 1
cell(ws3, rr, 20, "Margin", fill=GREENDK, font_name=GS, size=9, bold=True, color=WHITE, border=None, ha="left")
cell(ws3, rr, 21, 0.15, fill=GREENLT, font_name=AR, bold=True, border=None, ha="right", nf=PCT)
MARGIN = f"$U${rr}"

# ---- role columns (I..N) : (header seniority, header function, rate_cell) ----
roles = [
    ("Lead Consultant", "Project Manager", rate_cell["Lead Consultant"]),
    ("Principal", "Solution Architect / Tech Lead", rate_cell["Principal"]),
    ("Senior Consultant", "Backend / AI Engineer", rate_cell["Senior Consultant"]),
    ("Consultant", "Frontend Engineer", rate_cell["Consultant"]),
    ("Senior Consultant", "Data / BI Engineer", rate_cell["Senior Consultant"]),
    ("Consultant", "QA Engineer", rate_cell["Consultant"]),
]

# ---- phases: (milestone, activity, resources, weeks, [md per role], risk %) ----
ps_phases = [
    ("Phase 1 : GCP Foundation & Provisioning",
     "Provision GCP project, VPC, IAM, GKE Standard cluster + node pools, Cloud SQL, Memorystore, BigQuery, "
     "Artifact Registry; CI/CD, Secret Manager, monitoring baseline; re-platform Chatwoot + agent + backend "
     "from Docker Compose to GKE manifests.",
     "Solution Architect\nDevOps / Backend\nData/BI Engineer\nProject Manager", 3, [3, 10, 12, 0, 3, 2], 0.10),
    ("Phase 2 : Data Migration & Channel Cutover",
     "Zendesk extract, field mapping & bulk import (tickets, contacts, attachments); KB migration to pgvector + "
     "Vertex AI Search; re-wire WhatsApp/Website/API/Email inboxes on the new cluster; parallel-run & reconciliation.",
     "Backend/AI Engineer\nData/BI Engineer\nFrontend Engineer\nQA\nProject Manager", 4, [4, 4, 16, 4, 6, 6], 0.20),
    ("Phase 3 : Platform Features outside Packages A–G",
     "Native SAML SSO (IdP metadata, ACS endpoint, JIT provisioning, Security settings UI); RBAC phase 3 — native "
     "conversation-visibility scoping; IVR-4 voice-language reliability fix; custom agent status labels + routing "
     "integration; release & enablement (Chatwoot image rebuild with patches 0034–0041, RSA DB provisioning, "
     "BigQuery view rollout, per-tenant feature flags).",
     "Backend/AI Engineer\nFrontend Engineer\nSolution Architect\nQA\nData/BI Engineer", 8, [3, 9, 33, 13, 3, 15], 0.20),
    ("Phase 4 : Hardening, HA & UAT",
     "HA / autoscaling & resource tuning; monitoring, alerting & logging; backups and disaster-recovery runbook; "
     "live channel verification (real WhatsApp, phone call, browser smoke); UAT, data reconciliation & sign-off.",
     "DevOps / Backend\nQA\nSolution Architect\nProject Manager", 4, [6, 5, 11, 3, 2, 20], 0.10),
    ("Phase 5 : Cutover, Decommission & Handover",
     "Go-live cutover & hypercare; Zendesk decommission; ops runbook + admin/agent training & knowledge transfer.",
     "Project Manager\nBackend/AI Engineer\nDevOps\nTechnical Writer", 2, [6, 3, 5, 2, 2, 3], 0.05),
]

# title + headers
box(ws3, 3, 2, 3, 6, v="Professional Services — PROTON CRM · scope outside Packages A–G",
    fill=PINK, font_name=GS, bold=True, wrap=True, ha="center")
for c, h in [(2, "Milestone"), (3, "Activity"), (4, "Resources"), (5, "Duration (Week)"), (6, "Service Price")]:
    cell(ws3, 4, c, h, fill=PINK, font_name=GS, bold=True, ha="center", wrap=True)
cell(ws3, 4, 8, "Duration (Days)", fill=YELLOW, font_name=GS, bold=True, ha="center", wrap=True)
for i, (sen, fn, _) in enumerate(roles):
    cell(ws3, 3, 9 + i, sen, fill=YELLOW, font_name=GS, size=9, bold=True, ha="center", wrap=True)
    cell(ws3, 4, 9 + i, fn, fill=YELLOW, font_name=GS, size=9, bold=True, ha="center", wrap=True)
cell(ws3, 3, 15, "Total", fill=YELLOW, font_name=GS, bold=True, ha="center")
cell(ws3, 4, 15, "(mandays)", fill=YELLOW, font_name=GS, size=9, bold=True, ha="center")
box(ws3, 3, 16, 3, 18, v="Risk buffer", fill=GREY2, font_name=GS, bold=True, color=GREYTX, ha="center")
cell(ws3, 4, 16, "Risk %", fill=GREY2, font_name=GS, size=9, bold=True, ha="center", wrap=True)
cell(ws3, 4, 17, "Contingency (mandays)", fill=GREY2, font_name=GS, size=9, bold=True, ha="center", wrap=True)
cell(ws3, 4, 18, "Total incl. contingency", fill=GREY2, font_name=GS, size=9, bold=True, ha="center", wrap=True)

first = 5
r = first
for (mil, act, res, wks, mds, risk) in ps_phases:
    cell(ws3, r, 2, mil, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 3, act, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 4, res, font_name=GS, wrap=True, ha="left")
    cell(ws3, r, 5, wks, font_name=GS, ha="center", nf='#,##0')
    cell(ws3, r, 8, f"=E{r}*5", font_name=GS, ha="center")
    for i, md in enumerate(mds):
        cell(ws3, r, 9 + i, md, font_name=GS, ha="center")
    cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=GS, ha="center")
    cell(ws3, r, 16, risk, font_name=GS, ha="center", nf=PCT)
    cell(ws3, r, 17, f"=ROUND(O{r}*P{r},0)", font_name=GS, ha="center")
    cell(ws3, r, 18, f"=O{r}+Q{r}", font_name=GS, bold=True, ha="center")
    ws3.row_dimensions[r].height = 72
    r += 1
last = r - 1

# Service Price = one merged cell (F5:F9) = grand total incl. contingency & margin, rounded
box(ws3, first, 6, last, 6, v=f"=ROUNDUP(O{last+5},-2)", font_name=GS, bold=True, ha="center", nf=RP0_FMT)

# ---- totals block ----
cell(ws3, r, 8, "Total Mandays", fill=YELLOW, font_name=AR, bold=True, ha="left")
for i in range(len(roles)):
    cl = get_column_letter(9 + i)
    cell(ws3, r, 9 + i, f"=SUM({cl}{first}:{cl}{last})", font_name=AR, ha="center")
cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=AR, bold=True, ha="center")
cell(ws3, r, 16, None, fill=GREY2)
cell(ws3, r, 17, f"=SUM(Q{first}:Q{last})", fill=GREY2, font_name=AR, bold=True, ha="center")
cell(ws3, r, 18, f"=O{r}+Q{r}", fill=GREY2, font_name=AR, bold=True, ha="center")
row_md = r; r += 1

cell(ws3, r, 8, "Total Mandays Price", fill=YELLOW, font_name=AR, bold=True, ha="left")
for i, (_, _, ratec) in enumerate(roles):
    cl = get_column_letter(9 + i)
    cell(ws3, r, 9 + i, f"={cl}{row_md}*8*{ratec}", font_name=AR, ha="center", nf=RP0_FMT)
cell(ws3, r, 15, f"=SUM(I{r}:N{r})", font_name=AR, bold=True, ha="center", nf=RP0_FMT)
row_price = r; r += 1

cell(ws3, r, 8, "Contingency (risk buffer)", fill=YELLOW, font_name=AR, bold=True, ha="left", wrap=True)
box(ws3, r, 9, r, 14, v=None)
cell(ws3, r, 15, f"=O{row_price}/O{row_md}*Q{row_md}", font_name=AR, ha="center", nf=RP0_FMT)
row_cont = r; r += 1

cell(ws3, r, 8, "Subtotal incl. Contingency", fill=YELLOW, font_name=AR, bold=True, ha="left", wrap=True)
box(ws3, r, 9, r, 14, v=None)
cell(ws3, r, 15, f"=O{row_price}+O{row_cont}", font_name=AR, bold=True, ha="center", nf=RP0_FMT)
row_sub = r; r += 1

cell(ws3, r, 8, "Total Price Including Margin", fill=YELLOW, font_name=AR, bold=True, ha="left", wrap=True)
box(ws3, r, 9, r, 14, v=None)
cell(ws3, r, 15, f"=O{row_sub}*(1+{MARGIN})", font_name=AR, bold=True, ha="center", nf=RP0_FMT)
row_total = r; r += 2

# ---- memo band: packages A-G, excluded from every total above ----
box(ws3, r, 2, r, 6, v="Memo — excluded from the totals above (priced separately)",
    fill=GREY2, font_name=GS, bold=True, color=GREYTX, ha="left")
r += 1
cell(ws3, r, 2, "Packages A–G (in delivery)", font_name=GS, wrap=True, ha="left", color=GREYTX)
cell(ws3, r, 3, "A· Facebook/Instagram activation · B· Contacts ↔ Customer 360 merge · C· Telephony (live handoff, "
                "transcript-to-ticket, auto-busy, call recording) · D· Demo data & per-case detail · E· Reporting "
                "parity with the weekly/monthly decks · F· DMS/TSP integration · G· Escalation-policy engine.",
     font_name=GS, wrap=True, ha="left", color=GREYTX)
cell(ws3, r, 4, "Already queued —\nseparate estimate", font_name=GS, wrap=True, ha="left", color=GREYTX)
cell(ws3, r, 5, None)
cell(ws3, r, 6, "—", font_name=GS, ha="center", color=GREYTX)
ws3.row_dimensions[r].height = 60
r += 2

cell(ws3, r, 2, "Scope is zero-based on the work outside Packages A–G: platform delivery that has not started "
                "(GCP landing zone, GKE re-platform, Zendesk + KB migration, cutover, UAT, handover) plus the "
                "product items that have a spec/plan but no code (native SAML SSO, RBAC phase 3, IVR-4, custom "
                "agent status labels, release/enablement and live verification). Mandays are an indicative delivery "
                "estimate; day-rates, margin and per-phase Risk % are editable (rate card in T:V, Risk % in column P). "
                "May be offset by Google program funding.",
     border=None, font_name=AR, size=9, color="FF999999", ha="left", wrap=True)
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=15)
ws3.row_dimensions[r].height = 56
fit_landscape(ws3)

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
