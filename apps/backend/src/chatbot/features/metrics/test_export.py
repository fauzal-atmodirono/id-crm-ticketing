import asyncio
import io
from dataclasses import fields

from openpyxl import load_workbook  # type: ignore[import-untyped]

from chatbot.features.metrics.export import render_pdf, render_xlsx
from chatbot.features.metrics.query_port import DashboardMetrics, MockMetricsQuery


async def _metrics() -> DashboardMetrics:
    return await MockMetricsQuery().fetch_dashboard()


def _sync_metrics() -> DashboardMetrics:
    return asyncio.run(_metrics())


def test_render_xlsx_is_a_zip_workbook() -> None:
    data = render_xlsx(_sync_metrics())
    assert data[:2] == b"PK"  # xlsx is a zip container
    assert len(data) > 500


def test_render_xlsx_has_a_sheet_per_block() -> None:
    wb = load_workbook(io.BytesIO(render_xlsx(_sync_metrics())))
    for block in [f.name for f in fields(DashboardMetrics)]:
        assert block in wb.sheetnames


def test_render_pdf_has_pdf_header() -> None:
    data = render_pdf(_sync_metrics())
    assert data[:5] == b"%PDF-"
    assert len(data) > 500
